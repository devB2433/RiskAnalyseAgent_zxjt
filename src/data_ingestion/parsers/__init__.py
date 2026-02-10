"""
数据解析器实现
"""
import json
import csv
import xml.etree.ElementTree as ET
from typing import Any, Dict, List, Optional, Union
from io import StringIO

from ..base import BaseParser, DataFormat


class JSONParser(BaseParser):
    """JSON解析器"""

    def __init__(self):
        super().__init__(DataFormat.JSON)

    async def parse(
        self,
        raw_data: Union[str, bytes, Any],
        options: Optional[Dict[str, Any]] = None
    ) -> List[Dict[str, Any]]:
        """解析JSON数据"""
        if isinstance(raw_data, bytes):
            raw_data = raw_data.decode('utf-8')

        if isinstance(raw_data, str):
            data = json.loads(raw_data)
        else:
            data = raw_data

        # 确保返回列表
        if isinstance(data, dict):
            return [data]
        elif isinstance(data, list):
            return data
        else:
            return [{"value": data}]

    def validate_format(self, raw_data: Union[str, bytes, Any]) -> bool:
        """验证JSON格式"""
        try:
            if isinstance(raw_data, bytes):
                raw_data = raw_data.decode('utf-8')
            if isinstance(raw_data, str):
                json.loads(raw_data)
            return True
        except (json.JSONDecodeError, UnicodeDecodeError):
            return False


class CSVParser(BaseParser):
    """CSV解析器"""

    def __init__(self):
        super().__init__(DataFormat.CSV)

    async def parse(
        self,
        raw_data: Union[str, bytes, Any],
        options: Optional[Dict[str, Any]] = None
    ) -> List[Dict[str, Any]]:
        """解析CSV数据"""
        if isinstance(raw_data, bytes):
            raw_data = raw_data.decode('utf-8')

        options = options or {}
        delimiter = options.get('delimiter', ',')
        has_header = options.get('has_header', True)

        reader = csv.DictReader(
            StringIO(raw_data),
            delimiter=delimiter
        ) if has_header else csv.reader(StringIO(raw_data), delimiter=delimiter)

        result = []
        if has_header:
            for row in reader:
                result.append(dict(row))
        else:
            for i, row in enumerate(reader):
                result.append({f"col_{j}": val for j, val in enumerate(row)})

        return result

    def validate_format(self, raw_data: Union[str, bytes, Any]) -> bool:
        """验证CSV格式"""
        try:
            if isinstance(raw_data, bytes):
                raw_data = raw_data.decode('utf-8')
            # 简单验证：尝试读取第一行
            csv.reader(StringIO(raw_data))
            return True
        except Exception:
            return False


class XMLParser(BaseParser):
    """XML解析器"""

    def __init__(self):
        super().__init__(DataFormat.XML)

    async def parse(
        self,
        raw_data: Union[str, bytes, Any],
        options: Optional[Dict[str, Any]] = None
    ) -> List[Dict[str, Any]]:
        """解析XML数据"""
        if isinstance(raw_data, bytes):
            raw_data = raw_data.decode('utf-8')

        root = ET.fromstring(raw_data)
        return self._element_to_dict(root, options)

    def _element_to_dict(
        self,
        element: ET.Element,
        options: Optional[Dict[str, Any]] = None
    ) -> List[Dict[str, Any]]:
        """将XML元素转换为字典"""
        result = []

        def parse_element(elem) -> Dict[str, Any]:
            data = {
                'tag': elem.tag,
                'attributes': elem.attrib,
                'text': elem.text.strip() if elem.text else None
            }

            children = []
            for child in elem:
                children.append(parse_element(child))

            if children:
                data['children'] = children

            return data

        if options and options.get('root_as_list', False):
            for child in element:
                result.append(parse_element(child))
        else:
            result.append(parse_element(element))

        return result

    def validate_format(self, raw_data: Union[str, bytes, Any]) -> bool:
        """验证XML格式"""
        try:
            if isinstance(raw_data, bytes):
                raw_data = raw_data.decode('utf-8')
            ET.fromstring(raw_data)
            return True
        except ET.ParseError:
            return False


class ExcelParser(BaseParser):
    """Excel解析器"""

    def __init__(self):
        super().__init__(DataFormat.EXCEL)

    async def parse(
        self,
        raw_data: Union[str, bytes, Any],
        options: Optional[Dict[str, Any]] = None
    ) -> List[Dict[str, Any]]:
        """解析Excel数据"""
        try:
            import openpyxl
            from io import BytesIO

            if isinstance(raw_data, str):
                # 文件路径
                workbook = openpyxl.load_workbook(raw_data)
            else:
                # 字节数据
                workbook = openpyxl.load_workbook(BytesIO(raw_data))

            options = options or {}
            sheet_name = options.get('sheet_name', None)
            has_header = options.get('has_header', True)

            # 选择工作表
            if sheet_name:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            result = []
            rows = list(sheet.iter_rows(values_only=True))

            if not rows:
                return result

            if has_header:
                headers = rows[0]
                for row in rows[1:]:
                    result.append({
                        str(headers[i]): row[i]
                        for i in range(len(row))
                        if i < len(headers)
                    })
            else:
                for row in rows:
                    result.append({
                        f"col_{i}": val
                        for i, val in enumerate(row)
                    })

            return result

        except ImportError:
            raise ImportError("请安装 openpyxl 库: pip install openpyxl")

    def validate_format(self, raw_data: Union[str, bytes, Any]) -> bool:
        """验证Excel格式"""
        try:
            import openpyxl
            from io import BytesIO

            if isinstance(raw_data, str):
                openpyxl.load_workbook(raw_data)
            else:
                openpyxl.load_workbook(BytesIO(raw_data))
            return True
        except Exception:
            return False


class LogParser(BaseParser):
    """日志解析器（支持多种日志格式）"""

    def __init__(self, log_format: str = "syslog"):
        super().__init__(DataFormat.LOG)
        self.log_format = log_format

    async def parse(
        self,
        raw_data: Union[str, bytes, Any],
        options: Optional[Dict[str, Any]] = None
    ) -> List[Dict[str, Any]]:
        """解析日志数据"""
        if isinstance(raw_data, bytes):
            raw_data = raw_data.decode('utf-8')

        lines = raw_data.strip().split('\n')
        result = []

        for line in lines:
            if not line.strip():
                continue

            if self.log_format == "syslog":
                parsed = self._parse_syslog(line)
            elif self.log_format == "json":
                parsed = self._parse_json_log(line)
            elif self.log_format == "apache":
                parsed = self._parse_apache_log(line)
            else:
                parsed = {"raw": line}

            if parsed:
                result.append(parsed)

        return result

    def _parse_syslog(self, line: str) -> Optional[Dict[str, Any]]:
        """解析Syslog格式"""
        import re
        # 简化的Syslog解析
        pattern = r'<(\d+)>(\w+\s+\d+\s+\d+:\d+:\d+)\s+(\S+)\s+(\S+):\s+(.*)'
        match = re.match(pattern, line)

        if match:
            return {
                "priority": match.group(1),
                "timestamp": match.group(2),
                "hostname": match.group(3),
                "process": match.group(4),
                "message": match.group(5)
            }
        return {"raw": line}

    def _parse_json_log(self, line: str) -> Optional[Dict[str, Any]]:
        """解析JSON格式日志"""
        try:
            return json.loads(line)
        except json.JSONDecodeError:
            return {"raw": line}

    def _parse_apache_log(self, line: str) -> Optional[Dict[str, Any]]:
        """解析Apache日志格式"""
        import re
        # Apache Combined Log Format
        pattern = r'(\S+) \S+ \S+ \[(.*?)\] "(.*?)" (\d+) (\d+) "(.*?)" "(.*?)"'
        match = re.match(pattern, line)

        if match:
            return {
                "ip": match.group(1),
                "timestamp": match.group(2),
                "request": match.group(3),
                "status": int(match.group(4)),
                "size": int(match.group(5)),
                "referer": match.group(6),
                "user_agent": match.group(7)
            }
        return {"raw": line}

    def validate_format(self, raw_data: Union[str, bytes, Any]) -> bool:
        """验证日志格式"""
        # 日志格式比较宽松，基本都能解析
        return True


__all__ = [
    'JSONParser',
    'CSVParser',
    'XMLParser',
    'ExcelParser',
    'LogParser',
]